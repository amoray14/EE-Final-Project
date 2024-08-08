import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import os
import openpyxl
from openpyxl.utils import get_column_letter
import requests
import pandas as pd
import base64

def encode_image(image_path):
    """Encode the image to base64."""
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

def gpt_talk(jpg_path):
    """Interacts with GPT to analyze a JPEG image and updates global feedback_comments."""
    global feedback_comments, points_deducted

    # OpenAI API Key
    api_key = ""

    # Getting the base64 string
    base64_image = encode_image(jpg_path)

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}"
    }

    payload = {
        "model": "gpt-4-vision-preview",
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "q1: Locate the title 'Project Number:' in the image and provide the number that follows.return only the number in the form 'x-x-x-x' and not anything else. start the answer with the title question1: (Note that you keep the order of the answers)"
                    },
                    {
                        "type": "text",
                        "text": "q2: who is the advisor in the project?. start the answer with the title question2:"
                    },
                    {
                        "type": "text",
                        "text": "q3: Who is the presenter of the project? Usually it's under names. only if there is two names appear, separate them with the word and. start the answer with the title question3:"
                    },
                    {
                        "type": "text",
                        "text": "q4: what is the topic of the project? Usually its appear at the first sentence in the project.start the answer with the title question4:"
                    },
                    {
                        "type": "text",
                        "text": "q5: Is the background white? answer colorful or white.if  most of the background ia white answer white. If the background is white except for text or graphs, answer white. start the answer with the title question5:"
                    },
                    {
                        "type": "text",
                        "text": "q6: there is connection between the topic and the introduction? Return 0.6 if there is a complete match. 1.6 If there is a very good but not complete match. 2.6 If there is a partial match. 3.6 If there is no match at all start the answer with the title question6:"
                    },
                    {
                        "type": "text",
                        "text": "q7: there is connection between the introduction and the motivation? Return 0.7 if there is a complete match. 1.7 If there is a very good but not complete match. 2.7 If there is a partial match.3.7 If there is no match at all start the answer with the title question7:"
                    },
                    {
                        "type": "text",
                        "text": "q8: is the conclusions coonnected to the results? Return 0.8 if there is a complete match. 1.8 If there is a very good but not complete match. 2.8 If there is a partial match. 3.8 If there is no match at all  start the answer with the title question8:"
                    },
                    {
                        "type": "text",
                        "text": "q9: Read the poster from left to right. Column by column. Is the overall quality of the poster good? Return 0.9 if the poster is excellent. 1.9 If the poster is good but there are problems with it. 2.9 If the poster is at a low level. 3.9 The poster is bad. start the answer with the title question9:"
                    },
                    {
                        "type": "text",
                        "text": "q10: Summarize the poster up to 4 lines. Consider the connections between the different paragraphs, and assess the quality of the content. start the answer with the title question10:"
                    },
                    {
                        "type": "text",
                        "text": "q11: How relevant are the graphs to the project goals? How well are the graphs displayed in the results section? And how related are they to the verbal content? If you think the graphs are related to the content and are clear, return 0.11. If the graphs are related but not clear, return 1.11. If the graphs are unrelated, return 2.11. start the answer with the title question11:"
                    },
                    {
                        "type": "text",
                        "text": "q12: How well is the introduction written? And are the conclusions related to the introduction? If the introduction is excellent and there is a connection to the conclusions, return 0.12. If the introduction is excellent but there is no connection to the conclusions, return 1.12. If the introduction is good but needs to be improved and there is a connection to the conclusions, return 2.12. If the introduction is not good, return 3. start the answer with the title question12:"
                    },
                    {
                        "type": "text",
                        "text": "q13: How detailed is the implementation section? How well are the implementation steps described? If the section is specified, return 0.13. If the section is good but there missing details, return 1.13. If the section is unclear, return 2.13. start the answer with the title question13:"
                    },
                    {
                        "type": "text",
                        "text": "q14: Summarize the poster up to 4 lines. Consider the connections between the different paragraphs, and assess the quality of the content. start the answer with the title question14:"
                    },
                    {
                        "type": "text",
                        "text": "q15: Please read the poster column by column from left to right. Assess if the connections between the sections in the poster are strong and coherent. Provide your opinion in one line on the overall quality. Evaluate the quality of the explanations and the clarity of the entire poster. If the explanations are clear, conclude with the statement 'The section's explanations in the poster are clear'. If there is too much verbal data, conclude with 'The poster contains too much verbal information'. If there is a lack of visual explanation, conclude with 'Visual explanation is missing'. Additionally, evaluate the quality of the visual representation and the clarity of the entire poster. If the visuals of the poster are clear, conclude with the statement 'The poster visuality is good'. start the answer with the title question15:"
                    },

                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 600
    }

    response = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload)

    if response.status_code == 200:
        response_json = response.json()
        content = response_json['choices'][0]['message']['content']

        # Split the input string into an array of lines
        input_array = content.split('\n')

        # Initialize an empty array to store the extracted data
        feedback_comments = []

        for line in input_array:
            _, data = line.split(":", 1)  # Split each line by ":" to separate the question index from the data
            feedback_comments.append(data.strip())  # Add the extracted data to the array, stripping any leading or trailing whitespace

def update_feedback():
    global feedback_comments, points_deducted, filename, prompts
    if not feedback_comments:
        messagebox.showerror("Error", "No feedback data available.")
        return

    if not filename:
        messagebox.showerror("Error", "No image loaded.")
        return

    # Clear previous data in the table
    for row in feedback_table.get_children():
        feedback_table.delete(row)

    # Insert feedback comments, points deducted, and prompts into the table
    for idx, (comment, points) in enumerate(zip(feedback_comments[:len(feedback_comments)-2], points_deducted[:len(points_deducted)-2]), 1):
        if idx > 15:  # Stop after 15 items
            break
        feedback_table.insert("", "end", values=(idx, prompts[idx - 1], comment, points))  # Adjust index for prompts

    # Update the text boxes with the specific feedback comments
    update_text_boxes()

    excel_filename = os.path.join(os.path.dirname(filename), "projects.xlsx")
    wb = openpyxl.load_workbook(excel_filename)
    ws = wb.active

    # Get the image name without the extension
    image_name = os.path.basename(filename).replace('.jpg', '')

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            if cell.value == image_name:
                cell_offset = cell.offset(column=1)
                cell_offset.value = 100 - sum(points_deducted)

                # Insert project number in the new 'Project Number' column
                cell_offset.offset(column=2).value = feedback_comments[0]

                # Insert project summary in the new 'Project Summary' column
                cell_offset.offset(column=3).value = feedback_comments[13]

                # Insert evaluation summary in the new 'Evaluation Summary' column
                cell_offset.offset(column=4).value = feedback_comments[14]

                break
        else:
            continue
        break

    else:
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=4, value=image_name)
        ws.cell(row=next_row, column=5, value=100 - sum(points_deducted))

        # Insert project number in the new 'Project Number' column
        ws.cell(row=next_row, column=6).value = feedback_comments[0]

        # Insert project summary in the new 'Project Summary' column
        ws.cell(row=next_row, column=7).value = feedback_comments[13]

        # Insert evaluation summary in the new 'Evaluation Summary' column
        ws.cell(row=next_row, column=8).value = feedback_comments[14]

    # Add headers for the new columns if they don't exist
    if ws.cell(row=1, column=7).value is None:
        ws.cell(row=1, column=7, value="Project Summary")
    if ws.cell(row=1, column=8).value is None:
        ws.cell(row=1, column=8, value="Evaluation Summary")

    # Adjust the width of columns D, G, and H
    for col in ['D', 'G', 'H']:
        ws.column_dimensions[col].width = 50

    df = pd.DataFrame(ws.values)
    df.columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']  # Adjust column names
    df['E'] = pd.to_numeric(df['E'], errors='coerce')
    df = df.iloc[1:].copy()
    df.sort_values(by='E', ascending=False, inplace=True)

    for r_idx, row in enumerate(df.values, 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(excel_filename)
    messagebox.showinfo("Excel Updated",
                        f"Excel file {excel_filename} updated with the image name, project number, project summary, evaluation summary, and final grade, sorted by grade.")

def export_feedback():
    global feedback_comments, points_deducted, filename, prompts
    if not feedback_comments:
        messagebox.showerror("Error", "No feedback data available.")
        return

    if not filename:
        messagebox.showerror("Error", "No image loaded.")
        return

    project_number = feedback_comments[0]
    names = feedback_comments[2]
    excel_filename = f"{project_number}_{names}.xlsx"

    # Create a new Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add headers
    ws.append(["Index", "Prompt", "Feedback", "Points"])

    # Add feedback data
    for idx, (comment, points) in enumerate(zip(feedback_comments[:13], points_deducted[:13]), 1):  # Show only first 13 items
        ws.append([idx, prompts[idx - 1], comment, points])  # Adjust index for prompts

    # Add additional feedback comments to the secondary excel file
    ws.append(["", "Project Summary", feedback_comments[13], ""])
    ws.append(["", "Evaluation Summary", feedback_comments[14], ""])

    # Add the grade information
    final_grade = 100 - sum(points_deducted[:13])
    ws.append(["", "Grade", final_grade, ""])

    # Adjust the width of columns B and C
    for col in ['B', 'C']:
        ws.column_dimensions[col].width = 50

    # Save the Excel file
    wb.save(excel_filename)
    messagebox.showinfo("Export Successful", f"Feedback data exported to {excel_filename}.")

def calc():
    global feedback_comments, points_deducted
    for i in range(len(feedback_comments)):
        value = feedback_comments[i]
        try:
            num = float(value)
            if num.is_integer():
                points_deducted[i] = int(num)
            else:
                points_deducted[i] = int(num)  # Taking the lower bound
        except ValueError:
            continue

def update_text_boxes():
    text_box_9.config(state=tk.NORMAL)
    text_box_9.delete("1.0", tk.END)
    text_box_9.insert(tk.END, feedback_comments[13])
    text_box_9.config(state=tk.DISABLED)

    text_box_10.config(state=tk.NORMAL)
    text_box_10.delete("1.0", tk.END)
    text_box_10.insert(tk.END, feedback_comments[14])
    text_box_10.config(state=tk.DISABLED)

def load_and_analyze_image():
    global filename
    filename = filedialog.askopenfilename(title="Choose an image file", filetypes=(("JPEG files", "*.jpg"), ("All files", "*.*")))
    if filename:
        gpt_talk(filename)
        calc()
        update_feedback()
        export_feedback()

def main():
    global feedback_comments, filename, points_deducted, feedback_table, prompts, text_box_9, text_box_10
    points_deducted = [0] * 50
    feedback_comments = [""] * 50
    filename = ""
    prompts = [
        "Project Number", "Advisor", "Presenter", "Project Topic", "Background Color",
        "Topic-Introduction Connection", "Introduction-Motivation Connection", "Conclusions-Results Connection",
        "Overall Poster Quality", "Poster Summary", "Graphs Relevance and Clarity",
        "Introduction Quality and Connection to the Conclusions",
        "Implementation Quality"
    ]

    root = tk.Tk()
    root.title("Image Feedback Analysis")
    root.geometry("1000x800")  # Set the size of the window

    load_button = tk.Button(root, text="Load and Analyze Image", command=load_and_analyze_image)
    load_button.pack(pady=20)

    feedback_table = ttk.Treeview(root, columns=("Index", "Prompt", "Feedback", "Points"), show="headings")
    feedback_table.heading("Index", text="Index")
    feedback_table.heading("Prompt", text="Prompt")
    feedback_table.heading("Feedback", text="Feedback")
    feedback_table.heading("Points", text="Points Deducted")
    feedback_table.column("Index", width=20)  # Set the width of the Index column
    feedback_table.column("Points", width=20)  # Set the width of the Points column
    feedback_table.pack(padx=10, pady=10, expand=True, fill='both')

    # Create a frame to hold the text boxes
    text_frame = tk.Frame(root)
    text_frame.pack(pady=10, expand=True, fill='both')

    # Text box for feedback_comments[9]
    text_label_9 = tk.Label(text_frame, text="Project Summary:")
    text_label_9.pack(anchor="w")
    text_box_9 = tk.Text(text_frame, height=10, width=100, state=tk.DISABLED)
    text_box_9.pack(pady=5, expand=True, fill='both')

    # Text box for feedback_comments[10]
    text_label_10 = tk.Label(text_frame, text="Evaluation Summary:")
    text_label_10.pack(anchor="w")
    text_box_10 = tk.Text(text_frame, height=10, width=100, state=tk.DISABLED)
    text_box_10.pack(pady=5, expand=True, fill='both')

    root.mainloop()


if __name__ == "__main__":
    main()
